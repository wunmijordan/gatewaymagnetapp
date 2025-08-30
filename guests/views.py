from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, Http404
from .models import GuestEntry, FollowUpReport, SocialMediaEntry, Review
from .forms import GuestEntryForm, FollowUpReportForm
import csv
import io
from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model, authenticate, login
from django.core.paginator import Paginator
from django.contrib.auth.models import Group, User
from django.db.models import Q, Count, Max, F
from django.utils.http import urlencode
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib import messages
from datetime import datetime, timedelta
from django.utils import timezone
from django.utils.timezone import localtime, now, localdate
import pytz
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.template.loader import render_to_string, get_template
import weasyprint
#from .utils import get_week_start_end
from guests.models import GuestEntry
from django.utils.dateparse import parse_date
from django.db.models.functions import ExtractYear, ExtractMonth, TruncMonth
import calendar
import base64
import json
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
import os
from django.db import IntegrityError, transaction
from django.middleware.csrf import get_token
from urllib.parse import urlencode
from django.conf import settings
from cloudinary.uploader import upload as cloudinary_upload
from accounts.models import CustomUser
from urllib.parse import urlencode
from accounts.utils import user_in_groups





User = get_user_model()



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')  # Checkbox value

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            if remember_me:
                # Session will expire in 30 days
                request.session.set_expiry(60 * 60 * 24 * 30)
            else:
                # Session expires when browser closes
                request.session.set_expiry(0)

            return redirect('dashboard')
        else:
            context = {'form': LoginForm(request.POST)}
            return render(request, 'login.html', context)

    else:
        context = {'form': LoginForm()}
        return render(request, 'login.html', context)




@login_required
def dashboard_view(request):
    """Main dashboard view with server-rendered stats for cards and charts."""
    user = request.user
    current_year = datetime.now().year
    guest_entries = GuestEntry.objects.all()  # all guest entries for available years filter

    # Queryset for filtered data cards, charts (role based)
    if user_in_groups(request.user, "Pastor,Team Lead,Registrant,Admin"):
        queryset = GuestEntry.objects.all()
    else:
        queryset = GuestEntry.objects.filter(assigned_to=user)

    available_years = guest_entries.dates('date_of_visit', 'year')
    available_years = [d.year for d in available_years]


    # Pre-fill for current year summary
    request.GET = request.GET.copy()
    request.GET['year'] = str(current_year)
    summary_json = guest_entry_summary(request)
    summary_data = summary_json.content.decode()

    # Services attended (used for most attended card & chart)
    service_qs = GuestEntry.objects.values('service_attended').annotate(count=Count('id')).order_by('-count')
    service_labels = [s['service_attended'] or "Not Specified" for s in service_qs]
    service_counts = [s['count'] for s in service_qs]

    # Home Church stats
    total_purposes = GuestEntry.objects.count()
    home_church_count = GuestEntry.objects.filter(purpose_of_visit__iexact="Home Church").count()
    home_church_percentage = round((home_church_count / total_purposes) * 100, 1) if total_purposes else 0

    # Occasional Visit stats
    occasional_visit_count = GuestEntry.objects.filter(purpose_of_visit__iexact="Occasional Visit").count()
    occasional_visit_percentage = round((occasional_visit_count / total_purposes) * 100, 1) if total_purposes else 0

    # One-Time Visit stats
    one_time_visit_count = GuestEntry.objects.filter(purpose_of_visit__iexact="One-Time Visit").count()
    one_time_visit_percentage = round((one_time_visit_count / total_purposes) * 100, 1) if total_purposes else 0

    # Special Programme stats
    special_programme_count = GuestEntry.objects.filter(purpose_of_visit__iexact="Special Programme").count()
    special_programme_percentage = round((special_programme_count / total_purposes) * 100, 1) if total_purposes else 0

    # Most attended service card data
    if service_qs:
        most_attended_service = service_qs[0]['service_attended'] or "Not Specified"
        most_attended_count = service_qs[0]['count']
        total_services = sum(s['count'] for s in service_qs)
        attendance_rate = round((most_attended_count / total_services) * 100, 1) if total_services else 0
    else:
        most_attended_service = "No Data"
        most_attended_count = 0
        attendance_rate = 0

    # Status data (for cards)
    status_qs = GuestEntry.objects.values('status').annotate(count=Count('id'))
    status_labels = [s['status'] or "Unknown" for s in status_qs]
    status_counts = [s['count'] for s in status_qs]

    # Channel of Visit
    channel_qs = GuestEntry.objects.values('channel_of_visit').annotate(count=Count('id')).order_by('-count')
    total_channels = sum(c['count'] for c in channel_qs)
    channel_progress = [
        {
            'label': c['channel_of_visit'] or "Unknown",
            'count': c['count'],
            'percent': round((c['count'] / total_channels) * 100, 2) if total_channels else 0
        }
        for c in channel_qs
    ]

    # Totals & stats for cards (role filtered)
    planted_count = GuestEntry.objects.filter(status="Planted").count()
    planted_elsewhere_count = GuestEntry.objects.filter(status="Planted Elsewhere").count()
    relocated_count = GuestEntry.objects.filter(status="Relocated").count()
    work_in_progress_count = GuestEntry.objects.filter(status="Work in Progress").count()

    # === Global total guests and monthly increase rate (all users) ===
    total_guests = GuestEntry.objects.count()

    today = now().date()
    first_day_this_month = today.replace(day=1)
    last_month_end = first_day_this_month - timedelta(days=1)
    first_day_last_month = last_month_end.replace(day=1)

    current_month_count = GuestEntry.objects.filter(
        date_of_visit__gte=first_day_this_month,
        date_of_visit__lte=today
    ).count()

    last_month_count = GuestEntry.objects.filter(
        date_of_visit__gte=first_day_last_month,
        date_of_visit__lte=last_month_end
    ).count()

    if last_month_count == 0:
        if current_month_count > 0:
            increase_rate = 100
            percent_change = 100
        else:
            increase_rate = 0
            percent_change = 0
    else:
        difference = current_month_count - last_month_count
        increase_rate = round((difference / last_month_count) * 100, 1)
        percent_change = increase_rate

    # === Logged-in user's total guest entries and month difference ===
    user_total_guest_entries = GuestEntry.objects.filter(assigned_to=user).count()

    user_current_month_count = GuestEntry.objects.filter(
        assigned_to=user,
        date_of_visit__gte=first_day_this_month,
        date_of_visit__lte=today
    ).count()

    user_last_month_count = GuestEntry.objects.filter(
        assigned_to=user,
        date_of_visit__gte=first_day_last_month,
        date_of_visit__lte=last_month_end
    ).count()

    if user_last_month_count == 0:
        if user_current_month_count > 0:
            user_diff_percent = 100
            user_diff_positive = True
        else:
            user_diff_percent = 0
            user_diff_positive = True
    else:
        diff = ((user_current_month_count - user_last_month_count) / user_last_month_count) * 100
        user_diff_percent = round(abs(diff), 1)
        user_diff_positive = diff >= 0

    # === Planted guests growth rate for logged-in user ===
    user_planted_total = GuestEntry.objects.filter(assigned_to=user, status="Planted").count()
    planted_growth_rate = round((user_planted_total / user_total_guest_entries) * 100, 1) if user_total_guest_entries else 0

    user_planted_current_month = GuestEntry.objects.filter(
        assigned_to=user,
        status="Planted",
        date_of_visit__gte=first_day_this_month,
        date_of_visit__lte=today
    ).count()

    user_planted_last_month = GuestEntry.objects.filter(
        assigned_to=user,
        status="Planted",
        date_of_visit__gte=first_day_last_month,
        date_of_visit__lte=last_month_end
    ).count()

    if user_planted_last_month == 0:
        if user_planted_current_month > 0:
            planted_growth_change = 100
        else:
            planted_growth_change = 0
    else:
        diff = ((user_planted_current_month - user_planted_last_month) / user_planted_last_month) * 100
        planted_growth_change = round(diff, 1)

    # Load illustration image as base64
    #image_path = 'static/tabler/folders.png'
    #with open(image_path, 'rb') as img:
    #    image_data_uri = f"data:image/png;base64,{base64.b64encode(img.read()).decode()}"

    # Add all users except the logged-in user
    other_users = User.objects.exclude(id=request.user.id)

    context = {
        'show_filters': False,
        'available_years': available_years,
        'current_year': current_year,
        'summary_data': json.loads(summary_data),
        "service_labels": service_labels,
        "service_counts": service_counts,
        "status_labels": status_labels,
        "status_counts": status_counts,
        "channel_progress": channel_progress,
        "planted_count": planted_count,
        "planted_elsewhere_count": planted_elsewhere_count,
        "relocated_count": relocated_count,
        "work_in_progress_count": work_in_progress_count,
        "total_guests": total_guests,               # Global total guests
        "increase_rate": increase_rate,             # Global monthly increase rate (%)
        "percent_change": percent_change,           # Same as increase_rate for display
        "user_total_guest_entries": user_total_guest_entries,     # User's total guest entries
        "user_guest_entry_diff_percent": user_diff_percent,       # User's month-over-month % difference (absolute)
        "user_guest_entry_diff_positive": user_diff_positive,     # Boolean if user diff is positive
        "user_planted_total": user_planted_total,
        "planted_growth_rate": planted_growth_rate,                # User planted % of total user guests
        "planted_growth_change": planted_growth_change,            # User planted MoM % change
        #"image_data_uri": image_data_uri,
        "most_attended_service": most_attended_service,
        "most_attended_count": most_attended_count,
        "attendance_rate": attendance_rate,
        "home_church_count": home_church_count,
        "home_church_percentage": home_church_percentage,
        "occasional_visit_count": occasional_visit_count,
        "occasional_visit_percentage": occasional_visit_percentage,
        "one_time_visit_count": one_time_visit_count,
        "one_time_visit_percentage": one_time_visit_percentage,
        "special_programme_count": special_programme_count,
        "special_programme_percentage": special_programme_percentage,
        "other_users": other_users,
        "page_title": "Dashboard"
    }
    return render(request, "guests/dashboard.html", context)




def guest_entry_summary(request):
    year = request.GET.get('year')
    try:
        year = int(year)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid year'}, status=400)

    guests = GuestEntry.objects.filter(date_of_visit__year=year)

    # Total guests for the year
    total_count = guests.count()

    # Group by month and count
    month_counts = (
        guests.annotate(month=ExtractMonth('date_of_visit'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Build dict: {1: 5, 2: 10, ..., 12: 0}
    counts_dict = {month: 0 for month in range(1, 13)}
    for entry in month_counts:
        counts_dict[entry['month']] = entry['count']

    max_count = max(counts_dict.values())
    min_count = min(counts_dict.values())
    avg_count = sum(counts_dict.values()) // 12 if counts_dict else 0

    # Find month names
    max_months = [calendar.month_name[m] for m, c in counts_dict.items() if c == max_count]
    min_months = [calendar.month_name[m] for m, c in counts_dict.items() if c == min_count]

    # Just use the first if tie
    max_month = max_months[0] if max_months else "N/A"
    min_month = min_months[0] if min_months else "N/A"

    def percent(count):
        return round((count / max_count) * 100, 1) if max_count > 0 else 0

    data = {
        'max_month': max_month,
        'max_count': max_count,
        'max_percent': percent(max_count),
        'min_month': min_month,
        'min_count': min_count,
        'min_percent': percent(min_count),
        'avg_count': avg_count,
        'avg_percent': percent(avg_count),
        'total_count': total_count,
    }

    return JsonResponse(data)


def top_services_data(request):
    top_services = (
        GuestEntry.objects
        .values('service_attended')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Prepare data: list of dicts with service, count
    data = list(top_services)

    # Calculate total count of these top 10 to calculate % widths
    total = sum(item['count'] for item in data) or 1  # avoid division by zero

    # Add percentage to each
    for item in data:
        item['percent'] = round((item['count'] / total) * 100, 1)

    return JsonResponse({'services': data})




@login_required
def services_attended_chart(request):
    """AJAX endpoint for services attended chart."""

    queryset = GuestEntry.objects.all()  # No filtering by user

    qs = queryset.values('service_attended').annotate(count=Count('id')).order_by('-count')
    labels = [item['service_attended'] or "Not Specified" for item in qs]
    counts = [item['count'] for item in qs]

    return JsonResponse({'labels': labels, 'counts': counts})



@login_required
def channel_breakdown(request):
    """AJAX endpoint for channel of visit table."""
    
    queryset = GuestEntry.objects.all() 

    qs = queryset.values('channel_of_visit').annotate(count=Count('id')).order_by('-count')
    total = sum(item['count'] for item in qs)
    data = [
        {
            'label': item['channel_of_visit'] or 'Unknown',
            'count': item['count'],
            'percent': round((item['count'] / total) * 100, 2) if total else 0
        }
        for item in qs
    ]
    return JsonResponse(data, safe=False)




User = get_user_model()



@login_required
def guest_list_view(request):
    user = request.user
    role = user.username

    # --- GET filters ---
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    channel_filter = request.GET.get('channel', '').strip()
    purpose_filter = request.GET.get('purpose', '').strip()
    service_filter = request.GET.get('service', '').strip()
    user_filter = request.GET.get('user_filter', '').strip()
    date_of_visit_filter = request.GET.get('date_of_visit', '').strip()
    view_type = request.GET.get('view', 'cards')

    # --- Base queryset ---
    if user_in_groups(request.user, "Pastor,Team Lead,Registrant,Admin"):
        queryset = GuestEntry.objects.all()

    else:
        # Non-admins: show assigned guests OR the demo guest
        queryset = GuestEntry.objects.filter(
            Q(assigned_to=user) | Q(full_name="Wunmi Jordan")
        )

    # --- Apply search ---
    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query) |
            Q(service_attended__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(channel_of_visit__icontains=search_query) |
            Q(purpose_of_visit__icontains=search_query) |
            Q(assigned_to__full_name__icontains=search_query)
        )

    # --- Apply other filters ---
    if status_filter:
        queryset = queryset.filter(status__iexact=status_filter)
    if channel_filter:
        queryset = queryset.filter(channel_of_visit__iexact=channel_filter)
    if purpose_filter:
        queryset = queryset.filter(purpose_of_visit__iexact=purpose_filter)
    if service_filter:
        queryset = queryset.filter(service_attended__iexact=service_filter)
    if date_of_visit_filter:
        queryset = queryset.filter(date_of_visit=date_of_visit_filter)

    for guest in queryset:
        guest.has_unread_reviews = guest.reviews.filter(is_read=False).exists()

    # --- Annotate ---
    queryset = queryset.annotate(
        report_count=Count('reports'),
        last_reported=Max('reports__report_date')
    ).order_by('-custom_id')

    # --- Pagination ---
    per_page = 50 if view_type == 'list' else 12
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # --- Build query_string excluding 'page' for pagination links ---
    params = request.GET.copy()
    params.pop('page', None)
    query_string = urlencode(params)

    # --- Context ---
    context = {
        'page_obj': page_obj,
        'view_type': view_type,
        'users': User.objects.filter(is_active=True).order_by('first_name', 'last_name')[:100],
        'search_query': search_query,
        'status_filter': status_filter,
        'channel_filter': channel_filter,
        'purpose_filter': purpose_filter,
        'service_filter': service_filter,
        'user_filter': user_filter,
        'date_of_visit': date_of_visit_filter,
        'show_filters': True,
        'channels': GuestEntry.objects.values_list('channel_of_visit', flat=True).distinct().order_by('channel_of_visit'),
        'statuses': [s[0] for s in GuestEntry.STATUS_CHOICES],
        'purposes': GuestEntry.objects.values_list('purpose_of_visit', flat=True).distinct().order_by('purpose_of_visit'),
        'services': GuestEntry.objects.values_list('service_attended', flat=True).distinct().order_by('service_attended'),
        'query_string': query_string,
        'role': role,
        'page_title': 'Guests',
    }

    return render(request, 'guests/guest_list.html', context)




@user_passes_test(lambda u: user_in_groups(u, "Pastor,Team Lead,Registrant,Admin"))

@login_required
def create_guest(request):

    if request.method == 'POST':
        form = GuestEntryForm(request.POST, request.FILES)
        social_media_types = request.POST.getlist('social_media_type[]')
        social_media_handles = request.POST.getlist('social_media_handle[]')
        social_media_entries = []
        errors = []

        # Validate social media entries
        for i, (platform, handle) in enumerate(zip(social_media_types, social_media_handles)):
            platform = platform.strip()
            handle = handle.strip()
            if platform and handle:
                if platform not in dict(SocialMediaEntry.SOCIAL_MEDIA_CHOICES):
                    errors.append(f"Invalid social media platform at entry {i+1}.")
                elif len(handle) > 255:
                    errors.append(f"Handle too long at entry {i+1}.")
                else:
                    social_media_entries.append({'platform': platform, 'handle': handle})
            elif platform or handle:
                errors.append(f"Both platform and handle must be provided at entry {i+1}.")

        if form.is_valid() and not errors:
            guest = form.save(commit=False)
            guest.save()

            # Save social media
            for entry in social_media_entries:
                SocialMediaEntry.objects.create(guest=guest, **entry)

            # Handle redirect based on button clicked
            if 'save_add_another' in request.POST:
                return redirect('create_guest')  # stay on the form
            else:  # default Save button
                return redirect('guest_list')  # regular user guest list

        # If form invalid
        return render(request, 'guests/guest_form.html', {
            'form': form,
            'social_media_errors': errors,
            'edit_mode': False
        })

    else:
        form = GuestEntryForm()
        return render(request, 'guests/guest_form.html', {
            'form': form,
            'edit_mode': False,
            'page_title': 'Guests',
        })




@login_required
def edit_guest(request, pk):
    guest = get_object_or_404(GuestEntry, pk=pk)
    user = request.user

    # Permissions
    if guest.full_name == "Wunmi Jordan":
        # Allow everyone to edit this guest, but restrict certain actions
        pass
    elif not (user_in_groups(request.user, "Pastor,Team Lead,Admin") or guest.assigned_to == user):
        messages.error(request, "You do not have permission to edit this guest.")
        return redirect('guest_list')

    reassign_allowed = user_in_groups(request.user, "Pastor,Team Lead,Admin")
    all_users = User.objects.filter(is_active=True).order_by('full_name') if reassign_allowed else None
    social_media_entries = guest.social_media_accounts.all()

    if request.method == "POST":
        if "delete_guest" in request.POST:
            # Restrict deleting "Wunmi Jordan"
            if guest.full_name == "Wunmi Jordan" and not request.user.is_superuser:
                messages.error(request, "Only superusers can delete this guest.")
                return redirect("guest_list")

            guest.delete()
            messages.success(request, f"{guest.full_name} was deleted successfully.")
            return redirect("guest_list")

        form = GuestEntryForm(request.POST, request.FILES, instance=guest)

        # 🔒 Lock the name field in backend
        if guest.full_name == "Wunmi Jordan":
            form.fields["full_name"].disabled = True  # Prevent UI editing

        social_media_types = request.POST.getlist('social_media_type[]')
        social_media_handles = request.POST.getlist('social_media_handle[]')
        social_media_data = []
        errors = []

        for i, (platform, handle) in enumerate(zip(social_media_types, social_media_handles)):
            platform = platform.strip()
            handle = handle.strip()
            if platform and handle:
                if platform not in dict(SocialMediaEntry.SOCIAL_MEDIA_CHOICES):
                    errors.append(f"Invalid social media platform at entry {i+1}.")
                elif len(handle) > 255:
                    errors.append(f"Handle too long at entry {i+1}.")
                else:
                    social_media_data.append({'platform': platform, 'handle': handle})
            elif platform or handle:
                errors.append(f"Both platform and handle must be provided at entry {i+1}.")

        if form.is_valid() and not errors:
            updated_guest = form.save(commit=False)

            # 🔒 Ensure full_name remains unchanged
            if guest.full_name == "Wunmi Jordan":
                updated_guest.full_name = guest.full_name

            # Handle reassignment
            if reassign_allowed and 'assigned_to' in request.POST:
                assigned_id = request.POST.get('assigned_to')
                updated_guest.assigned_to = User.objects.filter(pk=assigned_id).first() if assigned_id else None

            # Clear picture if requested
            if 'clear_picture' in request.POST and guest.picture:
                guest.picture.delete(save=False)
                updated_guest.picture = None

            updated_guest.save()

            # Replace social media entries
            guest.social_media_accounts.all().delete()
            for entry in social_media_data:
                SocialMediaEntry.objects.create(guest=guest, **entry)

            # Redirect based on button clicked
            if 'save_add_another' in request.POST:
                return redirect('create_guest')  # Stay on new guest form
            else:  # default Save button
                return redirect('guest_list')

    else:
        form = GuestEntryForm(instance=guest)

        # 🔒 Lock the field in UI (read-only)
        if guest.full_name == "Wunmi Jordan":
            form.fields["full_name"].disabled = True

    return render(request, 'guests/guest_form.html', {
        'form': form,
        'guest': guest,
        'edit_mode': True,
        'can_reassign': reassign_allowed,
        'all_users': all_users,
        'show_delete': True,
        'social_media_entries': social_media_entries,
        'page_title': 'Guests',
    })




@login_required
def submit_review(request, guest_id, role):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    if request.method == "POST":
        comment = request.POST.get("comment")
        parent_id = request.POST.get("parent_id")
        parent = Review.objects.filter(id=parent_id).first() if parent_id else None
        Review.objects.create(
            guest=guest,
            reviewer=request.user,
            role=role,
            comment=comment,
            parent=parent
        )
    return redirect("guest_list")



@login_required
def mark_reviews_read(request, guest_id):
    guest = get_object_or_404(Guest, id=guest_id)
    
    # Only mark unread reviews for this user
    unread_reviews = guest.reviews.filter(is_read=False, reviewer=request.user)
    unread_reviews.update(is_read=True)
    
    return JsonResponse({"status": "success"})




def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)

@login_required
@superuser_required
def bulk_delete_guests(request):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        guest_ids = request.POST.getlist("guest_ids[]")
        if not guest_ids:
            return JsonResponse({"success": False, "message": "No guests selected."})
        
        deleted_count, _ = GuestEntry.objects.filter(id__in=guest_ids).delete()
        return JsonResponse({"success": True, "deleted_count": deleted_count})

    return JsonResponse({"success": False, "message": "Invalid request."})






User = get_user_model()

@login_required
def guest_detail_view(request, custom_id):
    """
    Returns rendered HTML for guest details,
    to be loaded directly into a Bootstrap modal.
    Bulletproof: handles missing social media, reports, and reassignment users.
    """

    # Define social media icons
    social_media_icons = {
        'linkedin': '''
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="none"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" 
            class="icon icon-tabler icon-tabler-brand-linkedin" style="color:#0A66C2;">
            <path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M8 11v5" />
            <path d="M8 8v.01" />
            <path d="M12 16v-5" />
            <path d="M16 16v-3a2 2 0 1 0 -4 0" />
            <path d="M3 7a4 4 0 0 1 4 -4h10a4 4 0 0 1 4 4v10a4 4 0 0 1 -4 4h-10a4 4 0 0 1 -4 -4z" />
        </svg>
        ''',
        'whatsapp': ''' 
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="none"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" 
            class="icon icon-tabler icon-tabler-brand-whatsapp" style="color:#25D366;">
            <path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M3 21l1.65 -3.8a9 9 0 1 1 3.4 2.9l-5.05 .9" />
            <path d="M9 10a.5 .5 0 0 0 1 0v-1a.5 .5 0 0 0 -1 0v1a5 5 0 0 0 5 5h1a.5 .5 0 0 0 0 -1h-1a.5 .5 0 0 0 0 1" />
        </svg>
        ''',
        'instagram': '''
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="none"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"
            class="icon icon-tabler icon-tabler-brand-instagram" style="color:#E4405F;">
            <path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <rect x="4" y="4" width="16" height="16" rx="4" />
            <circle cx="12" cy="12" r="3" />
            <line x1="16.5" y1="7.5" x2="16.5" y2="7.501" />
        </svg>
        ''',
        'twitter': '''
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="none"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"
            class="icon icon-tabler icon-tabler-brand-twitter" style="color:#1DA1F2;">
            <path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M22 4.01c-1 .49-1.98.689-3 .99c-1.121-1.265-2.783-1.335-4.38-.737
                    c-1.16 0-2.34.522-3.18 1.36c-1.208-.055-2.287-.616-3.07-1.52
                    c-.422.722-.666 1.561-.666 2.475c0 1.71.87 3.213 2.188 4.096
                    c-.807-.026-1.566-.247-2.229-.616c-.054 1.047.729 2.042 1.95 2.24
                    c-.693.188-1.452.232-2.224.084c.626 1.956 2.444 3.377 4.6 3.417
                    c-1.68 1.318-3.809 2.105-6.102 2.105c-.395 0-.779-.023-1.158-.067
                    c2.179 1.397 4.768 2.213 7.557 2.213c9.054 0 14-7.496 14-13.986
                    c0-.21 0-.423-.015-.633c.962-.689 1.8-1.56 2.46-2.548l-.047-.02z"/>
        </svg>
        ''',
        'tiktok': '''
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="none"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"
            class="icon icon-tabler icon-tabler-brand-tiktok" style="color:#000000;">
            <path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M9 19c-4 0-5-4-5-7a8 8 0 0 1 14-6v4" />
            <path d="M16 18.5a4.5 4.5 0 0 1-6.5-4" />
        </svg>
        ''',
    }

    user = request.user
    guest = get_object_or_404(GuestEntry, custom_id=custom_id)

    # Access control: allow only superusers/admins or creator/assigned user
    if not (user_in_groups(request.user, "Pastor,Team Lead,Admin")):
        if not (guest.assigned_to == user):
            return HttpResponse("Unauthorized", status=403)

    # Only admins and superusers get full user list for reassignment
    users = get_user_model().objects.all() if user_in_groups(request.user, "Pastor,Team Lead,Admin") else []

    # Fetch all social media accounts linked to this guest
    social_media_handles = guest.social_media_accounts.all()  # queryset, can be empty

    # Fetch all reports
    reports = guest.reports.all().order_by('-report_date') if hasattr(guest, 'reports') else []

    # Create form instance for display (readonly)
    form = GuestEntryForm(instance=guest)

    html = render_to_string('guests/guest_detail_modal.html', {
        'guest': guest,
        'social_media_icons': social_media_icons,
        'social_media_handles': social_media_handles,
        'reports': reports,
        'users': users,
        'form': form,
        'view_only': True,
        'page_title': f'Guest Detail - {guest.full_name}',
    }, request=request)

    return HttpResponse(html)





# -------------------------
# Reassign guest
# -------------------------
@login_required
@user_passes_test(lambda u: user_in_groups(u, "Pastor,Team Lead,Admin"))
def reassign_guest(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)

    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')

        if assigned_to_id:
            assigned_user = User.objects.filter(id=assigned_to_id, is_active=True).first()
            if assigned_user:
                guest.assigned_to = assigned_user
                guest.save()
                messages.success(request, f"Guest {guest.full_name} reassigned to {assigned_user.get_full_name() or assigned_user.username}.")
            else:
                messages.error(request, "Selected user does not exist or is inactive.")
        else:
            guest.assigned_to = None
            guest.save()
            messages.success(request, f"Assignment cleared for guest {guest.full_name}.")

    # Redirect admins back to admin dashboard
    return redirect('guest_list')






@require_POST
@login_required
def update_guest_status(request, pk):
    """
    Updates a guest's follow-up status (via dropdown in guest_list).
    Only the creator or an admin can update.
    """
    guest = get_object_or_404(GuestEntry, pk=pk)

    if not (user_in_groups(request.user, "Pastor,Team Lead,Admin") or guest.assigned_to == user):
        return redirect('guest_list')

    new_status = request.POST.get('status')
    if new_status in dict(GuestEntry.STATUS_CHOICES):
        guest.status = new_status
        guest.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))



def parse_flexible_date(date_str):
    """
    Try multiple date formats and return a valid `date` object or None.
    """
    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None



User = get_user_model()

@login_required
def import_guests_csv(request):
    if request.method != "POST" or not request.FILES.get("csv_file"):
        messages.error(request, "Please upload a valid CSV file.")
        return redirect("guest_list")

    csv_file = request.FILES["csv_file"]
    decoded_file = csv_file.read().decode("utf-8").splitlines()
    reader = csv.DictReader(decoded_file)

    guests_to_create = []

    # --- Step 1: Prepare GuestEntry objects ---
    for row in reader:
        username = row.get("assigned_to", "").strip()
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.warning(request, f"User '{username}' not found. Skipping row.")
            continue

        dob = row.get("date_of_birth", "").strip() or None
        dov = row.get("date_of_visit", "").strip() or None

        guest = GuestEntry(
            full_name=row.get("full_name", "").strip(),
            title=row.get("title", "").strip(),
            gender=row.get("gender", "").strip(),
            phone_number=row.get("phone_number", "").strip(),
            email=row.get("email", "").strip(),
            date_of_birth=dob,
            marital_status=row.get("marital_status", "").strip(),
            home_address=row.get("home_address", "").strip(),
            occupation=row.get("occupation", "").strip(),
            date_of_visit=dov,
            purpose_of_visit=row.get("purpose_of_visit", "").strip(),
            channel_of_visit=row.get("channel_of_visit", "").strip(),
            service_attended=row.get("service_attended", "").strip(),
            referrer_name=row.get("referrer_name", "").strip(),
            referrer_phone_number=row.get("referrer_phone_number", "").strip(),
            message=row.get("message", "").strip(),
            status=row.get("status", "").strip(),
            assigned_to=user,
            picture=row.get("picture_url", "").strip() or None  # <-- store Cloudinary URL directly
        )
        guests_to_create.append(guest)

    if not guests_to_create:
        messages.warning(request, "No valid guests found to import.")
        return redirect("guest_list")

    # --- Step 2: Bulk create guests ---
    with transaction.atomic():
        GuestEntry.objects.bulk_create(guests_to_create)

        # --- Step 3: Backfill custom_id ---
        prefix = "GNG"
        last_custom_id = GuestEntry.objects.filter(custom_id__startswith=prefix)\
            .aggregate(max_id=Max('custom_id'))['max_id']
        last_num = int(last_custom_id.replace(prefix, "")) if last_custom_id else 0

        new_guests = GuestEntry.objects.filter(custom_id__isnull=True).order_by("id")
        for idx, guest in enumerate(new_guests, start=last_num + 1):
            guest.custom_id = f"{prefix}{idx:06d}"

        GuestEntry.objects.bulk_update(new_guests, ["custom_id"])

    messages.success(request, f"{len(guests_to_create)} guests imported successfully!")
    return redirect("guest_list")






def download_csv_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="guest_import_template.csv"'

    writer = csv.writer(response)

    # Write header row
    writer.writerow([
        'full_name',               # Required
        'title',                   # Optional (Mr, Mrs, Miss, etc.)
        'gender',                  # Optional (Male, Female, Other)
        'phone_number',            # Optional
        'email',                   # Optional
        'date_of_birth',           # Optional (Any format like YYYY-MM-DD, DD/MM/YYYY, etc.)
        'marital_status',          # Optional
        'home_address',            # Optional
        'occupation',              # Optional
        'date_of_visit',           # Optional (Any format like YYYY-MM-DD, 24 July 2025, etc.)
        'purpose_of_visit',        # Optional
        'channel_of_visit',        # Optional (Flyer, Friend, Social Media, etc.)
        'service_attended',        # Optional (Sunday, Midweek, etc.)
        'referrer_name',           # Optional
        'referrer_phone_number',   # Optional
        'message',                 # Optional
        'status',                  # Optional (New, Returned, Not Interested, etc.)
        'assigned_to',              # Required (must match an existing username)
    ])

    # Optionally include one empty sample row
    writer.writerow([
        '', '', '', '', '',
        '', '', '', '',
        '', '', '', '',
        '', '', '', '', ''
    ])

    return response





@login_required
def export_csv(request):
    """
    Export filtered guest entries as CSV.
    - Admins can export all or filter by user.
    - Regular users can only export their own entries.
    - Respects service and search filters.
    """
    User = get_user_model()

    filter_user_id = request.GET.get('user')
    filter_service = request.GET.get('service')
    search_query = request.GET.get('q')

    # Base queryset
    if user_in_groups(request.user, "Pastor,Team Lead,Admin"):
        guests = GuestEntry.objects.all()
        if filter_user_id and filter_user_id.isdigit():
            guests = guests.filter(assigned_to__id=filter_user_id)
    else:
        guests = GuestEntry.objects.filter(assigned_to=request.user)

    # Service filter
    if filter_service:
        guests = guests.filter(service_attended__iexact=filter_service)

    # Search filter
    if search_query:
        guests = guests.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query)
        )

    # CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="guest_entries.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Full Name', 'Phone Number', 'Email', 'Gender',
        'Date of Birth', 'Marital Status', 'Home Address',
        'Occupation', 'Date of Visit', 'Purpose of Visit',
        'Channel of Visit', 'Service Attended',
        'Referrer Name', 'Referrer Phone Number', 'Status',
        'Assigned To'
    ])

    for guest in guests:
        writer.writerow([
            guest.full_name,
            guest.phone_number,
            guest.email,
            guest.gender,
            guest.date_of_birth,
            guest.marital_status,
            guest.home_address,
            guest.occupation,
            guest.date_of_visit,
            guest.purpose_of_visit,
            guest.channel_of_visit,
            guest.service_attended,
            guest.referrer_name,
            guest.referrer_phone_number,
            guest.status,
            guest.assigned_to.get_full_name() if guest.assigned_to else '',
        ])

    return response


@login_required
def update_status_view(request, guest_id, status_key):
    guest = get_object_or_404(GuestEntry, id=guest_id)

    # Only allow if user is the creator or admin
    if request.user != guest.assigned_to and not request.user.is_superuser:
        return redirect('guest_list')  # or return an HTTP 403 Forbidden response

    guest.status = status_key
    guest.save()
    return redirect('guest_list')




@login_required
def export_guests_excel(request):
    filter_user_id = request.GET.get('user')
    filter_service = request.GET.get('service')
    search_query = request.GET.get('q')

    guests = GuestEntry.objects.all() if user_in_groups(request.user, "Pastor,Team Lead,Admin") else GuestEntry.objects.filter(assigned_to=request.user)

    if filter_user_id and filter_user_id.isdigit():
        guests = guests.filter(assigned_to__id=filter_user_id)

    if filter_service:
        guests = guests.filter(service_attended__iexact=filter_service)

    if search_query:
        guests = guests.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query)
        )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guest Entries"

    headers = [
        'Full Name', 'Phone Number', 'Email', 'Gender',
        'Date of Birth', 'Marital Status', 'Home Address',
        'Occupation', 'Date of Visit', 'Purpose of Visit',
        'Channel of Visit', 'Service Attended',
        'Referrer Name', 'Referrer Phone Number', 'Status',
        'Assigned To'
    ]
    ws.append(headers)

    # Populate rows
    for guest in guests:
        ws.append([
            f"{guest.title} {guest.full_name}",
            guest.phone_number,
            guest.email,
            guest.gender,
            guest.date_of_birth.strftime('%Y-%m-%d') if guest.date_of_birth else "",
            guest.marital_status,
            guest.home_address,
            guest.occupation,
            guest.date_of_visit.strftime('%Y-%m-%d') if guest.date_of_visit else "",
            guest.purpose_of_visit,
            guest.channel_of_visit,
            guest.service_attended,
            guest.referrer_name,
            guest.referrer_phone_number,
            guest.status,
            guest.assigned_to.get_full_name() if guest.assigned_to else '',
        ])

    # Direct download as Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=guest_entries.xlsx'
    wb.save(response)
    return response





@login_required
def export_guests_pdf(request):
    guests = GuestEntry.objects.all()
    html = render_to_string('guests/guest_list_pdf.html', {'guests': guests})
    pdf_file = weasyprint.HTML(string=html).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'filename="all_guests.pdf"'
    return response




@login_required
@user_passes_test(lambda u: user_in_groups(u, "Pastor,Team Lead,Admin"))  # Only staff users can import
def import_guests_excel(request):
    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect("import_guests")

        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        username_col = headers.index("Assigned To (Username)")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                assigned_to_username = row[username_col]
                assigned_to_user = User.objects.get(username=assigned_to_username)

                GuestEntry.objects.create(
                    title=row[0],
                    full_name=row[1],
                    gender=row[2],
                    phone_number=row[3],
                    email=row[4],
                    date_of_birth=row[5],
                    marital_status=row[6],
                    home_address=row[7],
                    occupation=row[8],
                    date_of_visit=row[9],
                    purpose_of_visit=row[10],
                    channel_of_visit=row[11],
                    service_attended=row[12],
                    referrer_name=row[13],
                    referrer_phone_number=row[14],
                    message=row[15],
                    assigned_to=assigned_to_user,
                )

            except User.DoesNotExist:
                messages.warning(request, f"User '{assigned_to_username}' not found. Skipping row.")
                continue
            except Exception as e:
                messages.error(request, f"Error importing row: {e}")
                continue

        messages.success(request, "Guests imported successfully.")
        return redirect("guest_list")

    return render(request, "guests/import_excel.html")


def get_week_start_end(target_date):
    start = target_date - timedelta(days=target_date.weekday())
    end = start + timedelta(days=6)
    return start, end





@login_required
def followup_report_page(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    today = localdate()
    user = request.user

    # Permissions: admin or assigned user
    if not (user_in_groups(request.user, "Pastor,Team Lead,Admin") or guest.assigned_to == user):
        messages.error(request, "You do not have permission to edit this guest.")
        return redirect('guest_list')

    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    paginator = Paginator(reports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    error = None

    if request.method == 'POST' and 'submit_report' in request.POST:
        report_date = request.POST.get('date_of_visit') or localdate()
        note = request.POST.get('note')
        service_sunday = request.POST.get('service_sunday') == 'on'
        service_midweek = request.POST.get('service_midweek') == 'on'

        if not note:
            error = "Note field is required."
        else:
            try:
                FollowUpReport.objects.create(
                    guest=guest,
                    report_date=report_date,
                    note=note,
                    service_sunday=service_sunday,
                    service_midweek=service_midweek,
                    assigned_to=guest.assigned_to,  # Use the user assigned to the guest
                )
                return redirect('followup_report_page', guest_id=guest.id)
            except IntegrityError as e:
                if 'unique constraint' in str(e).lower() or 'duplicate key' in str(e).lower():
                    error = "Duplicate dates are not allowed."
                else:
                    raise

    return render(request, 'guests/followup_report_page.html', {
        'guest': guest,
        'reports': reports,
        'page_obj': page_obj,
        'today': today,
        'error': error,
    })


@login_required
def create_followup_report(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    user = request.user

    if request.method == "POST":
        form = FollowUpReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.guest = guest
            report.assigned_to = guest.assigned_to  # Assign the report to the guest's assigned user
            report.save()
            messages.success(request, "Follow-up report created successfully.")
            return redirect('guest_detail', guest_id=guest.id)
    else:
        form = FollowUpReportForm()

    return render(request, 'guests/followup_form.html', {
        'form': form,
        'guest': guest
    })





"""
def get_guest_reports(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    report_data = [
        {
            'report_date': localtime(report.report_date).strftime('%Y-%m-%d'),
            'note': report.note,
            'sunday_attended': report.sunday_attended,
            'midweek_attended': report.midweek_attended,
        }
        for report in reports
    ]
    return JsonResponse({'reports': report_data})
"""


@login_required
def followup_history_view(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    return render(request, 'guests/followup_history.html', {
        'guest': guest,
        'reports': reports,
    })


def export_followup_reports_pdf(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')

    # Create a PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Guest Reports")

    styles = getSampleStyleSheet()
    elements = []

    # Custom title style
    title_style = ParagraphStyle(
        name='Title',
        fontSize=16,
        leading=24,
        alignment=1,  # Center
        spaceAfter=20,
    )

    # Optional Logo
    logo_path = os.path.join('static', 'your_logo.png')  # Adjust path
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=100, height=40)
        logo.hAlign = 'LEFT'
        elements.append(logo)

    # Title
    elements.append(Paragraph(f"Follow-Up Report for {guest.full_name}", title_style))
    elements.append(Spacer(1, 10))

    # Table header and data
    data = [['Date', 'Sunday', 'Midweek', 'Message', 'Assigned To']]
    for report in reports:
        assigned_to_name = report.assigned_to.get_full_name() if report.assigned_to else 'Unknown'
        data.append([
            report.report_date.strftime("%Y-%m-%d"),
            '✔️' if report.service_sunday else '',
            '✔️' if report.service_midweek else '',
            report.note or ''
        ])

    table = Table(data, colWidths=[80, 60, 60, 280, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#000000")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="followup_reports_{guest.id}.pdf"'
    })



